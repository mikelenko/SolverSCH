import sys, io
sys.path.insert(0,'.')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl

wb = openpyxl.load_workbook('LLM_Test_Report.xlsx')
issues = []
ok_checks = []

# ────────────────────────────────────────────────────────────────
# 1. SUMMARY sheet
# ────────────────────────────────────────────────────────────────
ws = wb['Summary']
nodes = []
components = {}
reading_nodes = reading_components = False
for row in ws.iter_rows(values_only=True):
    if not any(v is not None for v in row):
        continue
    v0 = str(row[0]).strip() if row[0] else ''
    if v0 == 'Network Nodes': reading_nodes = True; reading_components = False; continue
    if v0 == 'Components Overview': reading_components = True; reading_nodes = False; continue
    if v0 == 'Reference': continue  # header
    if reading_nodes and row[1] is None and v0 not in ('Components Overview',):
        nodes.append(v0)
    if reading_components and row[1] is not None:
        components[v0] = {'type': row[1], 'value': row[2], 'terminals': row[3]}

print('=== SUMMARY ===')
print(f'  Nodes ({len(nodes)}): {nodes}')
print(f'  Components ({len(components)}):')
for ref, c in components.items():
    print(f'    {ref}: {c["type"]} = {c["value"]} | terminals: {c["terminals"]}')

# Physics check: gain from Rin and Rf
rin = rf = None
for ref, c in components.items():
    try:
        val = float(str(c['value']).split()[0])
        if 'Rin' in ref or ('in' in ref.lower() and c['type'] == 'Resistor'):
            rin = val
        if 'Rf' in ref or ref.startswith('Rf') or ref == 'Rf':
            rf = val
    except:
        pass

if rin and rf:
    gain_magnitude = rf / rin
    print(f'\n  [PHYSICS] Gain = Rf/Rin = {rf}/{rin} = {gain_magnitude:.2f}')
    if abs(gain_magnitude - 4.0) < 0.3:
        ok_checks.append(f'Gain = {gain_magnitude:.2f} ≈ 4 ✓')
    else:
        issues.append(f'BLAD WZMOCNIENIA: Rf/Rin = {gain_magnitude:.2f}, oczekiwane ~4.0')

# LP filter check: f_c = 1/(2*pi*R*C)
import math
r_filt = c_filt = None
for ref, c in components.items():
    try:
        val = float(str(c['value']).split()[0])
        if 'filter' in ref.lower() or 'lp' in ref.lower() or 'filt' in ref.lower():
            if c['type'] == 'Resistor':
                r_filt = val
            if c['type'] == 'Capacitor':
                c_filt = val
    except:
        pass

if r_filt and c_filt:
    f_cutoff = 1.0 / (2 * math.pi * r_filt * c_filt)
    print(f'  [PHYSICS] LP cutoff = 1/(2π·{r_filt}·{c_filt}) = {f_cutoff:.1f} Hz')
    if abs(f_cutoff - 159) < 20:
        ok_checks.append(f'LP cutoff = {f_cutoff:.1f} Hz ≈ 159 Hz ✓')
    else:
        issues.append(f'BLAD FILTRU: f_c = {f_cutoff:.1f} Hz (oczekiwane ~159 Hz)')

# ────────────────────────────────────────────────────────────────
# 2. DC Operating Point
# ────────────────────────────────────────────────────────────────
ws2 = wb['DC Operating Point']
dc_voltages = {}
for row in ws2.iter_rows(values_only=True):
    if row[0] and isinstance(row[0], str) and row[1] is not None and isinstance(row[1], (int, float)):
        dc_voltages[row[0]] = row[1]

print('\n=== DC Operating Point ===')
for node, v in dc_voltages.items():
    print(f'  V({node}) = {v} V')

# Check vcc
if dc_voltages.get('vcc') == 5.0:
    ok_checks.append('Vcc = 5.0V w DC ✓')
else:
    issues.append(f'BLAD DC: Vcc = {dc_voltages.get("vcc")} V (oczekiwane 5V)')

# Check: all signal nodes should be 0V at DC (AC source → 0V at DC)
for node in ('in', 'op_inv', 'op_out', 'out'):
    v = dc_voltages.get(node)
    if v is None:
        issues.append(f'BRAK w DC: wezel "{node}" nie ma wyniku')
    elif abs(v) > 0.01:
        issues.append(f'UWAGA DC: V({node}) = {v} V (oczekiwane 0V przy sterowaniu AC)')
    else:
        ok_checks.append(f'V({node}) = 0V w DC ✓')

# ────────────────────────────────────────────────────────────────
# 3. AC Sweep
# ────────────────────────────────────────────────────────────────
ws3 = wb['AC Sweep']
print(f'\n=== AC Sweep ({ws3.max_row} rows) ===')

# Find real header row (has "Freq" column)
header = None
data_start = None
for i, row in enumerate(ws3.iter_rows(values_only=True), 1):
    if row[0] and 'Freq' in str(row[0]):
        header = list(row)
        data_start = i + 1
        break

if header is None:
    issues.append('BLAD: Brak nagłówka w arkuszu AC Sweep')
else:
    print(f'  Header cols: {[h for h in header if h]}')

    # Find column indices
    def col(name):
        try: return header.index(name)
        except: return None

    col_freq = col('Freq [Hz]')
    col_in_mag  = col('|V(in)| [V]')
    col_out_mag = col('|V(out)| [V]')
    col_out_db  = col('V(out) [dB]')
    col_out_ph  = col('Phase(out) [°]')

    print(f'  Columns: freq={col_freq}, |V(in)|={col_in_mag}, |V(out)|={col_out_mag}, V(out)[dB]={col_out_db}')

    if col_out_db is None:
        issues.append('BLAD: Brak kolumny V(out)[dB] w AC Sweep — wezel "out" nie jest monitorowany')

    # Load AC data
    ac_rows = []
    for row in ws3.iter_rows(min_row=data_start, values_only=True):
        if row[col_freq] is not None and isinstance(row[col_freq], (int, float)):
            ac_rows.append(row)

    print(f'  Data rows: {len(ac_rows)}, freq range: {ac_rows[0][col_freq]:.1f} - {ac_rows[-1][col_freq]:.0f} Hz')

    if col_out_db is not None and ac_rows:
        # Mid-band gain (low freq, before LP cutoff)
        low_freq_rows = [r for r in ac_rows if r[col_freq] < 30]
        if low_freq_rows:
            r = low_freq_rows[0]
            v_in  = r[col_in_mag]  if col_in_mag  is not None else None
            v_out = r[col_out_mag] if col_out_mag is not None else None
            if v_in and v_out and v_in > 0:
                gain_ac = v_out / v_in
                print(f'  Mid-band gain @ {r[col_freq]:.1f} Hz: |Vout/Vin| = {gain_ac:.3f} (oczekiwane 4.0)')
                if abs(gain_ac - 4.0) < 0.5:
                    ok_checks.append(f'AC gain = {gain_ac:.2f} ≈ 4 ✓')
                else:
                    issues.append(f'BLAD GAIN AC: |Vout/Vin| = {gain_ac:.2f} przy low-freq (oczekiwane ~4.0)')

        # -3dB from LP filter at ~159Hz
        max_db = max(r[col_out_db] for r in ac_rows if r[col_out_db] is not None)
        crossings_3db = [(r[col_freq], r[col_out_db]) for r in ac_rows
                         if r[col_out_db] is not None and r[col_out_db] <= max_db - 3.0]
        if crossings_3db:
            f3db_measured = crossings_3db[0][0]
            print(f'  LP -3dB measured: {f3db_measured:.1f} Hz (oczekiwane ~159 Hz)')
            if abs(f3db_measured - 159) < 40:
                ok_checks.append(f'LP -3dB = {f3db_measured:.1f} Hz ≈ 159 Hz ✓')
            else:
                issues.append(f'BLAD FILTRU AC: -3dB @ {f3db_measured:.1f} Hz (oczekiwane ~159 Hz)')
        else:
            issues.append('UWAGA: Nie znaleziono punktu -3dB w zakresie pomiarowym AC')

        # Sample 3 key frequency points
        print('\n  Kluczowe punkty AC:')
        for f_target in (10, 100, 159, 1000, 10000):
            closest = min(ac_rows, key=lambda r: abs(r[col_freq] - f_target))
            db_v = closest[col_out_db]
            ph_v = closest[col_out_ph] if col_out_ph is not None else '?'
            print(f'    @ {f_target:>6} Hz : V(out) = {db_v:+.2f} dB, phase = {ph_v:.1f}°' if isinstance(ph_v, float) else f'    @ {f_target:>6} Hz : V(out) = {db_v:+.2f} dB')

# ────────────────────────────────────────────────────────────────
# 4. BOM
# ────────────────────────────────────────────────────────────────
ws4 = wb['BOM']
print(f'\n=== BOM ({ws4.max_row} rows) ===')
for row in ws4.iter_rows(values_only=True):
    if any(v is not None for v in row):
        fmt = [str(v)[:35] if v is not None else '' for v in row]
        print(f'  {fmt}')

# ────────────────────────────────────────────────────────────────
# FINAL VERDICT
# ────────────────────────────────────────────────────────────────
print('\n' + '='*60)
print(f'WYNIKI ANALIZY: {len(ok_checks)} OK, {len(issues)} problemów')
print('='*60)
print('\n✅ Poprawne:')
for o in ok_checks:
    print(f'  ✓ {o}')
if issues:
    print('\n⚠️  Problemy:')
    for iss in issues:
        print(f'  ✗ {iss}')
else:
    print('\nBrak błędów — raport wygląda poprawnie fizycznie!')
