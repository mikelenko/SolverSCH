"""
excel_report.py -> Excel Report Generator for SolverSCH Circuit Analysis.

Generates multi-sheet .xlsx workbooks with selectable analyses:
- Summary: Circuit topology overview
- DC Operating Point: Node voltages and source currents at DC
- AC Sweep: Frequency response with embedded Bode plot
- Transient: Time-domain waveforms with embedded chart
- BOM: Bill of Materials

Uses openpyxl for native Excel output with charts.
"""

import os
from typing import List, Optional, Dict, Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import NumericAxis
from openpyxl.utils import get_column_letter

from solver_sch.model.circuit import (
    Circuit, Component, Resistor, Capacitor, Inductor,
    VoltageSource, ACVoltageSource, Diode, BJT, MOSFET_N, MOSFET_P,
    OpAmp, Comparator
)
from solver_sch.builder.stamper import MNAStamper
from solver_sch.solver.sparse_solver import SparseSolver
import logging

logger = logging.getLogger("solver_sch.utils.excel_report")


# ── Styling Constants ──────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="4472C4")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

NUM_FORMAT_FREQ = '#,##0.00'
NUM_FORMAT_VOLTS = '0.000000'
NUM_FORMAT_DB = '0.00'
NUM_FORMAT_PHASE = '0.00'
NUM_FORMAT_TIME = '0.00E+00'


def _style_header_row(ws, row: int, col_count: int):
    """Apply professional header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 30):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min(max_len + 2, max_width), min_width)


def _component_type_name(comp: Component) -> str:
    """Returns a human-readable type name for a component."""
    type_map = {
        Resistor: "Resistor",
        Capacitor: "Capacitor",
        Inductor: "Inductor",
        VoltageSource: "DC Voltage Source",
        ACVoltageSource: "AC Voltage Source",
        Diode: "Diode",
        BJT: "BJT (NPN)",
        MOSFET_N: "NMOS FET",
        MOSFET_P: "PMOS FET",
        OpAmp: "Operational Amplifier",
        Comparator: "Comparator",
    }
    return type_map.get(type(comp), type(comp).__name__)


def _component_value_str(comp: Component) -> str:
    """Returns a formatted value string for a component."""
    if isinstance(comp, Resistor):
        return f"{comp.resistance} Ω"
    elif isinstance(comp, Capacitor):
        return f"{comp.capacitance} F"
    elif isinstance(comp, Inductor):
        return f"{comp.inductance} H"
    elif isinstance(comp, ACVoltageSource):
        return f"AC: {comp.ac_mag}V, {comp.amplitude}Vp @ {comp.frequency}Hz"
    elif isinstance(comp, VoltageSource):
        return f"{comp.voltage} V"
    elif isinstance(comp, OpAmp):
        return f"Gain: {comp.gain}"
    elif isinstance(comp, Diode):
        vz_str = f", Vz={comp.Vz}V" if comp.Vz else ""
        return f"Is={comp.Is}{vz_str}"
    return str(comp.value)


def _component_nodes_str(comp: Component) -> str:
    """Returns a string listing the component's terminal nodes."""
    return ", ".join(comp.nodes())


class ExcelReportGenerator:
    """Generates professional multi-sheet Excel reports from SolverSCH circuit analysis.
    
    Usage:
        report = ExcelReportGenerator(circuit)
        report.generate("output/report.xlsx", analyses=["summary", "dc", "ac", "transient", "bom"])
    """

    VALID_ANALYSES = {"summary", "dc", "ac", "transient", "bom"}

    def __init__(self, circuit: Circuit, ltspice_results: dict = None) -> None:
        self.circuit = circuit
        self._stamper: Optional[MNAStamper] = None
        self._solver: Optional[SparseSolver] = None
        self.ltspice_results = ltspice_results or {}

    def _ensure_solver(self):
        """Lazily build the MNA stamper and solver when first needed."""
        if self._stamper is None:
            self._stamper = MNAStamper(self.circuit)
            A_lil, z_vec = self._stamper.stamp_linear()
            self._solver = SparseSolver(
                A_lil, z_vec,
                self._stamper.node_to_idx,
                self._stamper.vsrc_to_idx,
                self._stamper.n
            )
            # Wire up callbacks for nonlinear + transient + AC
            self._solver.set_nonlinear_stamper(self._stamper.stamp_nonlinear)
            self._solver.set_ac_stamper(self._stamper.stamp_ac)
            self._solver.set_transient_stampers(
                self._stamper.stamp_transient_basis,
                self._stamper.stamp_transient_sources,
                self._stamper.update_states
            )
            self._solver.set_dynamic_stamper(self._stamper.stamp_dynamic_sources)

    def generate(
        self,
        filepath: str,
        analyses: Optional[List[str]] = None,
        ac_params: Optional[Dict[str, Any]] = None,
        transient_params: Optional[Dict[str, Any]] = None,
        auto_open: bool = True,
    ) -> str:
        """Generate the Excel report.

        Args:
            filepath: Output .xlsx file path.
            analyses: List of analyses to include. Options: "summary", "dc", "ac", "transient", "bom".
                      Defaults to all analyses.
            ac_params: Dict with keys "f_start", "f_stop", "ppd" (points per decade).
                       Defaults: f_start=100, f_stop=100e3, ppd=10.
            transient_params: Dict with keys "t_stop" (seconds), "dt" (seconds).
                              Defaults: t_stop=5e-3, dt=10e-6.
            auto_open: Whether to open the file in Excel after generation.

        Returns:
            Absolute path to the generated file.
        """
        if analyses is None:
            analyses = ["summary", "dc", "ac", "transient", "bom"]

        # Validate
        for a in analyses:
            if a not in self.VALID_ANALYSES:
                raise ValueError(f"Unknown analysis '{a}'. Valid: {self.VALID_ANALYSES}")

        # Defaults
        ac_p = {"f_start": 100, "f_stop": 100e3, "ppd": 10}
        if ac_params:
            ac_p.update(ac_params)

        trans_p = {"t_stop": 5e-3, "dt": 10e-6}
        if transient_params:
            trans_p.update(transient_params)

        # Create workbook
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        if "summary" in analyses:
            self._write_summary(wb)

        if "dc" in analyses:
            self._ensure_solver()
            self._write_dc(wb)

        if "ac" in analyses:
            self._ensure_solver()
            self._write_ac(wb, ac_p)

        if "transient" in analyses:
            self._ensure_solver()
            self._write_transient(wb, trans_p)
            
        if self.ltspice_results:
            self._write_ltspice_comparison(wb)

        if "bom" in analyses:
            self._write_bom(wb)

        # Save
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        abs_path = os.path.abspath(filepath)
        wb.save(abs_path)
        logger.info("[EXCEL] Report saved to %s", abs_path)

        if auto_open:
            try:
                os.startfile(abs_path)
                logger.debug("[EXCEL] Opening in Excel...")
            except Exception:
                logger.warning("[EXCEL] Open manually: %s", abs_path)

        return abs_path

    # ── LTspice Comparison ─────────────────────────────────────────

    def _write_ltspice_comparison(self, wb: Workbook):
        ws = wb.create_sheet("LTspice Comparison")

        ws.merge_cells("A1:F1")
        ws["A1"] = "SolverSCH vs LTspice Validation"
        ws["A1"].font = TITLE_FONT
        
        row_idx = 3

        for analysis_name, comp_result in self.ltspice_results.items():
            ws.cell(row=row_idx, column=1, value=f"{analysis_name.upper()} Analysis")
            ws.cell(row=row_idx, column=1).font = SUBTITLE_FONT
            
            status_text = "PASSED" if comp_result.passed else "FAILED"
            ws.cell(row=row_idx, column=2, value=status_text)
            ws.cell(row=row_idx, column=2).font = Font(bold=True, color="00B050" if comp_result.passed else "FF0000")
            
            row_idx += 1
            
            headers = ["Node/Metric", "Info", "SolverSCH", "LTspice", "Error [%]", "Status"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=h)
            _style_header_row(ws, row_idx, len(headers))
            
            row_idx += 1
            start_table = row_idx
            
            for nc in comp_result.nodes:
                ws.cell(row=row_idx, column=1, value=nc.node)
                ws.cell(row=row_idx, column=2, value=nc.info)
                
                ws.cell(row=row_idx, column=3, value=nc.solver_value)
                ws.cell(row=row_idx, column=3).number_format = "0.00E+00"
                
                ws.cell(row=row_idx, column=4, value=nc.ltspice_value)
                ws.cell(row=row_idx, column=4).number_format = "0.00E+00"
                
                ws.cell(row=row_idx, column=5, value=nc.error_pct)
                ws.cell(row=row_idx, column=5).number_format = "0.00"
                
                status_cell = ws.cell(row=row_idx, column=6, value=nc.status)
                status_cell.font = Font(bold=True)
                if nc.status == "PASS":
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100")
                elif nc.status == "WARN":
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    status_cell.font = Font(color="9C5700")
                else: # FAIL
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006")
                    
                row_idx += 1
                
            row_idx += 2 # space before next table
            
        _auto_width(ws)


    # ── Summary Sheet ──────────────────────────────────────────────

    def _write_summary(self, wb: Workbook):
        ws = wb.create_sheet("Summary")

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Circuit Analysis Report: {self.circuit.name}"
        ws["A1"].font = TITLE_FONT

        ws["A3"] = "Ground Node:"
        ws["A3"].font = SUBTITLE_FONT
        ws["B3"] = self.circuit.ground_name

        # Nodes
        ws["A5"] = "Network Nodes"
        ws["A5"].font = SUBTITLE_FONT
        nodes = sorted(self.circuit.get_unique_nodes())
        for i, node in enumerate(nodes):
            ws.cell(row=6 + i, column=1, value=node)

        # Components table
        row_start = 6 + len(nodes) + 2
        ws.cell(row=row_start, column=1, value="Components Overview")
        ws.cell(row=row_start, column=1).font = SUBTITLE_FONT

        headers = ["Reference", "Type", "Value", "Terminals"]
        hr = row_start + 1
        for col, h in enumerate(headers, 1):
            ws.cell(row=hr, column=col, value=h)
        _style_header_row(ws, hr, len(headers))

        for i, comp in enumerate(self.circuit.get_components()):
            r = hr + 1 + i
            ws.cell(row=r, column=1, value=comp.name)
            ws.cell(row=r, column=2, value=_component_type_name(comp))
            ws.cell(row=r, column=3, value=_component_value_str(comp))
            ws.cell(row=r, column=4, value=_component_nodes_str(comp))

        _auto_width(ws)

    # ── DC Operating Point ─────────────────────────────────────────

    def _write_dc(self, wb: Workbook):
        ws = wb.create_sheet("DC Operating Point")

        ws.merge_cells("A1:C1")
        ws["A1"] = "DC Operating Point Analysis"
        ws["A1"].font = TITLE_FONT

        # Solve DC
        result = self._solver.solve()

        # Node Voltages
        ws["A3"] = "Node Voltages"
        ws["A3"].font = SUBTITLE_FONT

        headers = ["Node", "Voltage [V]"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=h)
        _style_header_row(ws, 4, len(headers))

        row = 5
        for node in sorted(result.node_voltages.keys()):
            ws.cell(row=row, column=1, value=node)
            ws.cell(row=row, column=2, value=result.node_voltages[node])
            ws.cell(row=row, column=2).number_format = NUM_FORMAT_VOLTS
            row += 1

        # Source Currents
        row += 1
        ws.cell(row=row, column=1, value="Voltage Source Currents")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        row += 1

        headers = ["Source", "Current [A]"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        _style_header_row(ws, row, len(headers))
        row += 1

        for src in sorted(result.voltage_source_currents.keys()):
            ws.cell(row=row, column=1, value=src)
            ws.cell(row=row, column=2, value=result.voltage_source_currents[src])
            ws.cell(row=row, column=2).number_format = NUM_FORMAT_VOLTS
            row += 1

        _auto_width(ws)

    # ── AC Sweep + Bode Plot ───────────────────────────────────────

    def _write_ac(self, wb: Workbook, params: dict):
        ws = wb.create_sheet("AC Sweep")

        ws.merge_cells("A1:F1")
        ws["A1"] = "AC Frequency Sweep Analysis"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Range: {params['f_start']:.0f} Hz – {params['f_stop']:.0f} Hz | {params['ppd']} pts/decade"
        ws["A2"].font = Font(italic=True, color="666666")

        # Generate frequency points
        freqs = np.logspace(
            np.log10(params["f_start"]),
            np.log10(params["f_stop"]),
            int(np.log10(params["f_stop"] / params["f_start"]) * params["ppd"]) + 1
        )

        # Run AC analysis
        results = self._solver.simulate_ac(f_start=freqs.tolist(), stamper_ref=self._stamper)

        # Determine nodes (skip ground)
        _, first_mna = results[0]
        node_names = sorted([n for n in first_mna.node_voltages.keys() if n != self.circuit.ground_name])

        # Build header
        headers = ["Frequency [Hz]"]
        for node in node_names:
            headers.extend([f"|V({node})| [V]", f"V({node}) [dB]", f"Phase({node}) [°]"])

        hr = 4
        for col, h in enumerate(headers, 1):
            ws.cell(row=hr, column=col, value=h)
        _style_header_row(ws, hr, len(headers))

        # Write data
        for i, (freq, mna_res) in enumerate(results):
            r = hr + 1 + i
            ws.cell(row=r, column=1, value=freq)
            ws.cell(row=r, column=1).number_format = NUM_FORMAT_FREQ

            col = 2
            for node in node_names:
                v_complex = mna_res.node_voltages.get(node, 0)
                mag = abs(v_complex)
                db = 20 * np.log10(max(mag, 1e-20))
                phase = np.degrees(np.angle(v_complex))

                ws.cell(row=r, column=col, value=mag)
                ws.cell(row=r, column=col).number_format = NUM_FORMAT_VOLTS
                col += 1

                ws.cell(row=r, column=col, value=db)
                ws.cell(row=r, column=col).number_format = NUM_FORMAT_DB
                col += 1

                ws.cell(row=r, column=col, value=phase)
                ws.cell(row=r, column=col).number_format = NUM_FORMAT_PHASE
                col += 1

        data_rows = len(results)

        # ── Bode Plot (Magnitude dB) ──
        chart_mag = LineChart()
        chart_mag.title = "Bode Plot — Magnitude"
        chart_mag.style = 10
        chart_mag.y_axis.title = "Magnitude [dB]"
        chart_mag.x_axis.title = "Frequency [Hz]"
        chart_mag.width = 28
        chart_mag.height = 14

        freq_ref = Reference(ws, min_col=1, min_row=hr, max_row=hr + data_rows)

        for idx, node in enumerate(node_names):
            db_col = 2 + idx * 3 + 1  # dB column
            data_ref = Reference(ws, min_col=db_col, min_row=hr, max_row=hr + data_rows)
            chart_mag.add_data(data_ref, titles_from_data=True)
            chart_mag.set_categories(freq_ref)

        chart_mag.x_axis.scaling.logBase = 10
        chart_mag.x_axis.numFmt = '0'
        ws.add_chart(chart_mag, f"A{hr + data_rows + 3}")

        # ── Bode Plot (Phase) ──
        chart_phase = LineChart()
        chart_phase.title = "Bode Plot — Phase"
        chart_phase.style = 10
        chart_phase.y_axis.title = "Phase [°]"
        chart_phase.x_axis.title = "Frequency [Hz]"
        chart_phase.width = 28
        chart_phase.height = 14

        for idx, node in enumerate(node_names):
            phase_col = 2 + idx * 3 + 2  # Phase column
            data_ref = Reference(ws, min_col=phase_col, min_row=hr, max_row=hr + data_rows)
            chart_phase.add_data(data_ref, titles_from_data=True)
            chart_phase.set_categories(freq_ref)

        chart_phase.x_axis.scaling.logBase = 10
        chart_phase.x_axis.numFmt = '0'
        ws.add_chart(chart_phase, f"A{hr + data_rows + 20}")

        _auto_width(ws)

    # ── Transient Analysis + Chart ─────────────────────────────────

    def _write_transient(self, wb: Workbook, params: dict):
        ws = wb.create_sheet("Transient")

        ws.merge_cells("A1:F1")
        ws["A1"] = "Transient (Time-Domain) Analysis"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Duration: {params['t_stop']*1e3:.2f} ms | Step: {params['dt']*1e6:.1f} µs"
        ws["A2"].font = Font(italic=True, color="666666")

        # Run transient simulation
        results = self._solver.simulate_transient(params["t_stop"], params["dt"])

        if not results:
            ws["A4"] = "No transient results (empty simulation)."
            return

        # Determine nodes
        _, first_mna = results[0]
        node_names = sorted([n for n in first_mna.node_voltages.keys() if n != self.circuit.ground_name])

        # Header
        headers = ["Time [s]"] + [f"V({n}) [V]" for n in node_names]
        hr = 4
        for col, h in enumerate(headers, 1):
            ws.cell(row=hr, column=col, value=h)
        _style_header_row(ws, hr, len(headers))

        # Write data (downsample if too many points for Excel chart performance)
        max_rows = 2000
        step = max(1, len(results) // max_rows)
        sampled = results[::step]

        for i, (t, mna_res) in enumerate(sampled):
            r = hr + 1 + i
            ws.cell(row=r, column=1, value=t)
            ws.cell(row=r, column=1).number_format = NUM_FORMAT_TIME

            for j, node in enumerate(node_names):
                v = mna_res.node_voltages.get(node, 0.0)
                ws.cell(row=r, column=2 + j, value=v)
                ws.cell(row=r, column=2 + j).number_format = NUM_FORMAT_VOLTS

        data_rows = len(sampled)

        # ── Transient Waveform Chart ──
        chart = LineChart()
        chart.title = "Transient Waveforms"
        chart.style = 10
        chart.y_axis.title = "Voltage [V]"
        chart.x_axis.title = "Time [s]"
        chart.width = 28
        chart.height = 14

        time_ref = Reference(ws, min_col=1, min_row=hr, max_row=hr + data_rows)

        for j in range(len(node_names)):
            data_ref = Reference(ws, min_col=2 + j, min_row=hr, max_row=hr + data_rows)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(time_ref)

        ws.add_chart(chart, f"A{hr + data_rows + 3}")
        _auto_width(ws)

    # ── LTspice Comparison Sheet ─────────────────────────────────

    def _write_ltspice_comparison(self, wb: Workbook):
        """Write a cross-validation comparison sheet with PASS/WARN/FAIL formatting."""
        ws = wb.create_sheet("LTspice Comparison")

        PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        PASS_FONT = Font(color="006100", bold=True)
        WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        WARN_FONT = Font(color="9C6500", bold=True)
        FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        FAIL_FONT = Font(color="9C0006", bold=True)

        ws.merge_cells("A1:F1")
        ws["A1"] = "LTspice Cross-Validation Report"
        ws["A1"].font = TITLE_FONT

        current_row = 3

        for analysis_name, comparison in self.ltspice_results.items():
            # Section header
            ws.cell(row=current_row, column=1, value=f"{analysis_name.upper()} Analysis")
            ws.cell(row=current_row, column=1).font = SUBTITLE_FONT

            overall_status = "✅ PASSED" if comparison.passed else "❌ FAILED"
            ws.cell(row=current_row, column=3, value=f"Overall: {overall_status}")
            ws.cell(row=current_row, column=5, value=f"Tolerance: {comparison.tolerance_pct}%")
            current_row += 1

            # Table header
            headers = ["Node", "SolverSCH", "LTspice", "Error %", "Status", "Info"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=current_row, column=col, value=h)
            _style_header_row(ws, current_row, len(headers))
            current_row += 1

            # Data rows
            for nc in comparison.nodes:
                ws.cell(row=current_row, column=1, value=nc.node)
                ws.cell(row=current_row, column=2, value=round(nc.solver_value, 6) if not (nc.solver_value != nc.solver_value) else "N/A")
                ws.cell(row=current_row, column=3, value=round(nc.ltspice_value, 6) if not (nc.ltspice_value != nc.ltspice_value) else "N/A")
                ws.cell(row=current_row, column=4, value=round(nc.error_pct, 3) if not (nc.error_pct != nc.error_pct) else "N/A")

                status_cell = ws.cell(row=current_row, column=5, value=nc.status)
                if nc.status == "PASS":
                    status_cell.fill = PASS_FILL
                    status_cell.font = PASS_FONT
                elif nc.status == "WARN":
                    status_cell.fill = WARN_FILL
                    status_cell.font = WARN_FONT
                else:
                    status_cell.fill = FAIL_FILL
                    status_cell.font = FAIL_FONT

                ws.cell(row=current_row, column=6, value=nc.info)

                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).border = THIN_BORDER
                current_row += 1

            # Summary row
            ws.cell(row=current_row, column=1, value=f"Max Error: {comparison.max_error_pct:.3f}%")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 2  # blank row between analyses

        _auto_width(ws)

    # ── BOM Sheet ──────────────────────────────────────────────────

    def _write_bom(self, wb: Workbook):
        ws = wb.create_sheet("BOM")

        ws.merge_cells("A1:D1")
        ws["A1"] = "Bill of Materials"
        ws["A1"].font = TITLE_FONT

        headers = ["#", "Reference", "Type", "Value", "Terminals"]
        hr = 3
        for col, h in enumerate(headers, 1):
            ws.cell(row=hr, column=col, value=h)
        _style_header_row(ws, hr, len(headers))

        for i, comp in enumerate(self.circuit.get_components()):
            r = hr + 1 + i
            ws.cell(row=r, column=1, value=i + 1)
            ws.cell(row=r, column=2, value=comp.name)
            ws.cell(row=r, column=3, value=_component_type_name(comp))
            ws.cell(row=r, column=4, value=_component_value_str(comp))
            ws.cell(row=r, column=5, value=_component_nodes_str(comp))

        # Total
        total_row = hr + 1 + len(self.circuit.get_components()) + 1
        ws.cell(row=total_row, column=1, value="Total:")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=f"{len(self.circuit.get_components())} components")

        _auto_width(ws)
